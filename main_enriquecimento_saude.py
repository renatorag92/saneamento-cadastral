# main_enriquecimento_saude.py
# ==============================================
# ENRIQUECIMENTO COM BASE DA SAÚDE
# Lê o resultado da v2 e complementa dados das pessoas
# ==============================================

import pandas as pd
import os
from datetime import datetime
from unidecode import unidecode
import re

# ==============================================
# CONFIGURAÇÃO
# ==============================================
DATA_EXECUCAO = datetime.now().strftime('%Y%m%d_%H%M%S')
PASTA_OUTPUT = f'outputs/ENRIQUECIDO_{DATA_EXECUCAO}'
os.makedirs(PASTA_OUTPUT, exist_ok=True)

print("=" * 60)
print("🏁 ENRIQUECIMENTO COM BASE DA SAÚDE")
print(f"   Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
print("=" * 60)

# ==============================================
# FUNÇÕES
# ==============================================
def normalizar_doc(doc):
    """Remove máscara, trata float, garante 11 dígitos para CPF."""
    if pd.isna(doc):
        return ""
    
    limpo = re.sub(r'[^0-9]', '', str(doc))
    
    # Remove o ".0" do float (ex: "66863961415.0" → "66863961415")
    if limpo.endswith('0') and len(limpo) == 12:
        limpo = limpo[:-1]
    
    # Se tem 11 dígitos, é CPF válido
    if len(limpo) == 11:
        return limpo
    
    # Se tem 14 dígitos, pode ser CNPJ (mantém)
    if len(limpo) == 14:
        return limpo
    
    # CNS (15 dígitos) ou outros formatos — não é CPF
    return ""

def normalizar_nome(nome):
    if pd.isna(nome): return ""
    nome = unidecode(str(nome)).upper()
    nome = re.sub(r'[^A-Z\s]', '', nome)
    return re.sub(r'\s+', ' ', nome).strip()

def normalizar_telefone(tel):
    if pd.isna(tel) or str(tel) in ['-', '', 'nan']: return ""
    return re.sub(r'[^0-9]', '', str(tel))

# ==============================================
# CARREGAR BASES
# ==============================================
print("\n📂 Carregando bases...")

# Resultado da v2 (ajuste o caminho se necessário)
arquivo_v2 = 'outputs/CONSOLIDADO_v2_*/BASE_FINAL_COMPLETA.xlsx'
import glob
arquivos_v2 = glob.glob(arquivo_v2)
if not arquivos_v2:
    print("❌ Arquivo da v2 não encontrado. Rode a v2 primeiro.")
    exit()

arquivo_v2 = sorted(arquivos_v2)[-1]  # Pega o mais recente
print(f"   V2: {arquivo_v2}")

df_resultado = pd.read_excel(arquivo_v2)
df_saude = pd.read_excel('bases_brutas/base_saude_unificada.xlsx')

print(f"   Resultado v2: {len(df_resultado)} imóveis")
print(f"   Saúde: {len(df_saude)} cidadãos")

# ==============================================
# PADRONIZAR
# ==============================================
print("\n🧹 Padronizando...")

# Padroniza CPF no resultado (pode vir em formatos diferentes)
df_resultado['CPF_LIMPO'] = df_resultado['CPF_CNPJ_IPTU'].apply(normalizar_doc)

# Padroniza Saúde
df_saude['CPF_LIMPO'] = df_saude['CPF/CNS'].apply(normalizar_doc)
df_saude['NOME_PADRAO'] = df_saude['Nome'].apply(normalizar_nome)
df_saude['TELEFONE_LIMPO'] = df_saude['Telefone celular'].apply(normalizar_telefone)
df_saude['DATA_NASC_LIMPO'] = df_saude['Data de nascimento'].apply(
    lambda x: str(x).replace('/', '')[:8] if pd.notna(x) else ''
)

# Filtra só CPFs válidos (11 dígitos) na saúde
df_saude['CPF_VALIDO'] = df_saude['CPF_LIMPO'].apply(lambda x: x if len(x) == 11 else '')

# Cria índice por CPF para busca rápida
saude_por_cpf = {}
for _, row in df_saude.iterrows():
    cpf = row['CPF_VALIDO']
    if cpf:
        if cpf not in saude_por_cpf:
            saude_por_cpf[cpf] = []
        saude_por_cpf[cpf].append(row)

cpfs_unicos = len(saude_por_cpf)
print(f"   CPFs únicos na saúde: {cpfs_unicos}")
print("   ✅ Padronização concluída")

# ==============================================
# ENRIQUECIMENTO
# ==============================================
print("\n🔗 Enriquecendo com dados da saúde...")

# Novas colunas
df_resultado['SAUDE_NOME'] = ''
df_resultado['SAUDE_DATA_NASC'] = ''
df_resultado['SAUDE_TELEFONE'] = ''
df_resultado['SAUDE_CARTAO_SUS'] = ''
df_resultado['SAUDE_STATUS'] = 'NAO_ENCONTRADO'

encontrados = 0
multiplos = 0

for idx, row in df_resultado.iterrows():
    cpf = row['CPF_LIMPO']
    
    if cpf and len(cpf) == 11 and cpf in saude_por_cpf:
        matches = saude_por_cpf[cpf]
        
        if len(matches) == 1:
            m = matches[0]
            df_resultado.at[idx, 'SAUDE_NOME'] = m['NOME_PADRAO']
            df_resultado.at[idx, 'SAUDE_DATA_NASC'] = m['DATA_NASC_LIMPO']
            df_resultado.at[idx, 'SAUDE_TELEFONE'] = m['TELEFONE_LIMPO']
            df_resultado.at[idx, 'SAUDE_CARTAO_SUS'] = m['CPF/CNS']
            df_resultado.at[idx, 'SAUDE_STATUS'] = 'ENCONTRADO'
            encontrados += 1
        else:
            # Múltiplos registros para o mesmo CPF (pega o primeiro)
            m = matches[0]
            df_resultado.at[idx, 'SAUDE_NOME'] = m['NOME_PADRAO']
            df_resultado.at[idx, 'SAUDE_DATA_NASC'] = m['DATA_NASC_LIMPO']
            df_resultado.at[idx, 'SAUDE_TELEFONE'] = m['TELEFONE_LIMPO']
            df_resultado.at[idx, 'SAUDE_CARTAO_SUS'] = f"{len(matches)} registros"
            df_resultado.at[idx, 'SAUDE_STATUS'] = 'MULTIPLOS'
            multiplos += 1

print(f"   ✅ Encontrados: {encontrados}")
print(f"   ⚠️ Múltiplos registros: {multiplos}")
print(f"   ❌ Não encontrados: {len(df_resultado) - encontrados - multiplos}")

# ==============================================
# COMPARAÇÃO DE NOMES (qualidade)
# ==============================================
print("\n📊 Qualidade dos nomes...")
from fuzzywuzzy import fuzz

mudancas_nome = 0
for idx, row in df_resultado.iterrows():
    if row['SAUDE_STATUS'] in ['ENCONTRADO', 'MULTIPLOS']:
        nome_iptu = str(row.get('NOME_IPTU', ''))
        nome_saude = str(row['SAUDE_NOME'])
        if nome_iptu and nome_saude and nome_iptu != nome_saude:
            mudancas_nome += 1

print(f"   Nomes divergentes entre IPTU e Saúde: {mudancas_nome}")

# ==============================================
# EXPORTAR
# ==============================================
print("\n💾 Exportando...")

colunas_exportar = list(df_resultado.columns)
df_resultado[colunas_exportar].to_excel(f'{PASTA_OUTPUT}/BASE_ENRIQUECIDA_COMPLETA.xlsx', index=False)

# Indicadores
pd.DataFrame({
    'Indicador': [
        'Total de imóveis',
        'CPFs encontrados na saúde',
        'CPFs com múltiplos registros',
        'CPFs não encontrados',
        'Nomes divergentes (IPTU vs Saúde)',
        'Fonte',
        'Data'
    ],
    'Valor': [
        len(df_resultado),
        f"{encontrados} ({round(encontrados/len(df_resultado)*100,1)}%)",
        f"{multiplos} ({round(multiplos/len(df_resultado)*100,1)}%)",
        f"{len(df_resultado)-encontrados-multiplos}",
        mudancas_nome,
        'Base Saúde (22434 cidadãos)',
        datetime.now().strftime('%d/%m/%Y %H:%M')
    ]
}).to_excel(f'{PASTA_OUTPUT}/INDICADORES_ENRIQUECIMENTO.xlsx', index=False)

# ==============================================
# EXPORTAR COM ABAS SEPARADAS
# ==============================================
print("\n📂 Gerando arquivo com abas separadas...")

with pd.ExcelWriter(f'{PASTA_OUTPUT}/BASE_ENRIQUECIDA_CONSOLIDADA.xlsx') as writer:
    df_resultado[df_resultado['STATUS'] == 'CONFIRMADO'][colunas_exportar].to_excel(
        writer, sheet_name='CONFIRMADOS', index=False)
    df_resultado[df_resultado['STATUS'] == 'EM_REVISÃO'][colunas_exportar].to_excel(
        writer, sheet_name='EM_REVISAO', index=False)
    df_resultado[df_resultado['STATUS'] == 'PENDENTE'][colunas_exportar].to_excel(
        writer, sheet_name='PENDENTES', index=False)

# Contagem por status
for status in ['CONFIRMADO', 'EM_REVISÃO', 'PENDENTE']:
    qtd = len(df_resultado[df_resultado['STATUS'] == status])
    enriquecidos = len(df_resultado[(df_resultado['STATUS'] == status) & (df_resultado['SAUDE_STATUS'] == 'ENCONTRADO')])
    print(f"   {status}: {qtd} imóveis ({enriquecidos} enriquecidos)")

print(f"\n📁 Resultados em: {PASTA_OUTPUT}/")
print("   ✅ BASE_ENRIQUECIDA_COMPLETA.xlsx (tudo junto)")
print("   ✅ BASE_ENRIQUECIDA_CONSOLIDADA.xlsx (separado por abas)")
print("   ✅ INDICADORES_ENRIQUECIMENTO.xlsx")

# ==============================================
# PLANILHA DE REVISÃO ENRIQUECIDA (COM ENDEREÇO DA SAÚDE)
# ==============================================
print("\n📊 Gerando planilha de revisão enriquecida...")

df_revisao = df_resultado[df_resultado['STATUS'] == 'EM_REVISÃO'].copy()

import numpy as np

# Garante que ambos os CPFs sejam string, MAS substitui vazios por NaN para evitar o produto cartesiano
df_revisao['CPF_CNPJ_IPTU'] = df_revisao['CPF_CNPJ_IPTU'].fillna('').astype(str).str.replace('.0', '').str.strip().replace('', np.nan)
df_saude['CPF_VALIDO'] = df_saude['CPF_VALIDO'].fillna('').astype(str).str.strip().replace('', np.nan)

# Limpa duplicatas da saúde ANTES do merge (garante que um CPF retorne apenas 1 endereço)
df_saude_unicos = df_saude[['CPF_VALIDO', 'Endereço']].dropna(subset=['CPF_VALIDO']).drop_duplicates(subset='CPF_VALIDO', keep='first')

# Faz merge com a base da saúde para pegar o endereço bruto
df_revisao = df_revisao.merge(
    df_saude_unicos,
    left_on='CPF_CNPJ_IPTU',
    right_on='CPF_VALIDO',
    how='left'
)

# Remove duplicatas de segurança
antes = len(df_revisao)
df_revisao = df_revisao.drop_duplicates(subset=['OBJECTID', 'CPF_CNPJ_IPTU'], keep='first')
depois = len(df_revisao)
print(f"   Removidas {antes - depois} duplicatas do merge com saúde")

# Extrai componentes do endereço de TODAS as fontes
def extrair_parte(end, indice):
    if pd.isna(end) or not end:
        return ''
    partes = str(end).split(' | ')
    return partes[indice] if len(partes) > indice else ''

# Pesquisa e IPTU (formato canônico: LOGRADOUROADOURO | NUMEROERO | BAIRRORRO)
for fonte, col_orig in [('PESQUISA', 'ENDERECO_PESQUISA'), ('IPTU', 'ENDERECO_IPTU')]:
    for i, nome in [(0, 'LOGRADOURO'), (1, 'NUMERO'), (2, 'BAIRRO')]:
        df_revisao[f'{nome}_{fonte}'] = df_revisao[col_orig].apply(lambda x: extrair_parte(x, i))

# Saúde (formato bruto)
def extrair_LOGRADOUROadouro_saude(end):
    if pd.isna(end) or not end:
        return ''
    partes = str(end).split(',')
    return partes[0].strip() if len(partes) > 0 else str(end).strip()

def extrair_BAIRROrro_saude(end):
    if pd.isna(end) or not end:
        return ''
    partes = str(end).split(',')
    if len(partes) >= 3:
        return partes[1].strip().lstrip('0123456789. ')
    elif len(partes) >= 2:
        return partes[1].strip()
    return ''

df_revisao['LOGRADOURO_SAUDE'] = df_revisao['Endereço'].apply(extrair_LOGRADOUROadouro_saude)
df_revisao['BAIRRO_SAUDE'] = df_revisao['Endereço'].apply(extrair_BAIRROrro_saude)
df_revisao['NUMERO_SAUDE'] = ''  # Saúde não tem número separado

# Renomeia colunas para exibição limpa
renomeios = {
    'SAUDE_NOME': 'NOME_SAUDE',
    'SAUDE_CARTAO_SUS': 'CPF_SAUDE',
    'SAUDE_TELEFONE': 'TEL_SAUDE',
    'SAUDE_DATA_NASC': 'DATA_NASC',
    'CPF_CNPJ_PESQUISA': 'CPF_PESQUISA',
    'CPF_CNPJ_IPTU': 'CPF_IPTU',
    'TELEFONE_PESQUISA': 'TEL_PESQUISA',
    'TELEFONE_IPTU': 'TEL_IPTU',
}
df_revisao = df_revisao.rename(columns=renomeios)

# ==============================================
# REGRAS DE SOBREVIVÊNCIA E LAYOUT DE REVISÃO
# ==============================================
print("\n🤖 Calculando melhores fontes (regras de sobrevivência)...")

def escolher_fonte_nome(row):
    opcoes = {
        'SAUDE': str(row.get('NOME_SAUDE', '')).strip(),
        'IPTU': str(row.get('NOME_IPTU', '')).strip(),
        'PESQUISA': str(row.get('NOME_PESQUISA', '')).strip()
    }
    validas = {k: v for k, v in opcoes.items() if v not in ['nan', 'None', '', 'NAN']}
    if not validas: return 'NENHUM'
    return max(validas, key=lambda k: len(validas[k]))

def escolher_fonte_cpf(row):
    opcoes = {
        'SAUDE': str(row.get('CPF_SAUDE', '')).strip(),
        'IPTU': str(row.get('CPF_IPTU', '')).strip(),
        'PESQUISA': str(row.get('CPF_PESQUISA', '')).strip()
    }
    validas = {}
    for fonte, cpf in opcoes.items():
        limpo = re.sub(r'[^0-9]', '', cpf)
        if limpo and set(limpo) != {'0'}:
            if len(limpo) in [11, 14]: 
                return fonte 
            validas[fonte] = limpo
    
    if validas: return max(validas, key=lambda k: len(validas[k]))
    return 'NENHUM'

def escolher_fonte_telefone(row):
    opcoes = {
        'SAUDE': str(row.get('TEL_SAUDE', '')).strip(),
        'PESQUISA': str(row.get('TEL_PESQUISA', '')).strip(),
        'IPTU': str(row.get('TEL_IPTU', '')).strip()
    }
    validas = []
    for fonte, tel in opcoes.items():
        limpo = re.sub(r'[^0-9]', '', tel)
        if len(limpo) >= 10: return fonte
        if tel not in ['nan', 'None', '']: validas.append(fonte)
    return validas[0] if validas else 'NENHUM'

# 1. Preenche as colunas de FONTE
df_revisao['FONTE_NOME'] = df_revisao.apply(escolher_fonte_nome, axis=1)
df_revisao['FONTE_CPF'] = df_revisao.apply(escolher_fonte_cpf, axis=1)
df_revisao['FONTE_TEL'] = df_revisao.apply(escolher_fonte_telefone, axis=1)

# 2. Cria as colunas FINAL vazias
df_revisao['NOME_FINAL'] = ''
df_revisao['CPF_FINAL'] = ''
df_revisao['TEL_FINAL'] = ''

# 3. Layout sem a coluna REVISOR
colunas_organizadas = [
    'OBJECTID',
    'NOME_PESQUISA', 'NOME_IPTU', 'NOME_SAUDE',
    'CPF_PESQUISA', 'CPF_IPTU', 'CPF_SAUDE',
    'TEL_PESQUISA', 'TEL_IPTU', 'TEL_SAUDE',
    'LOGRADOURO_PESQUISA', 'LOGRADOURO_IPTU', 'LOGRADOURO_SAUDE',
    'NUMERO_PESQUISA', 'NUMERO_IPTU', 'NUMERO_SAUDE',
    'BAIRRO_PESQUISA', 'BAIRRO_IPTU', 'BAIRRO_SAUDE',
    'SCORE', 'MOTIVO', 'DATA_NASC', 'SAUDE_STATUS',
    'FONTE_NOME', 'NOME_FINAL',
    'FONTE_CPF', 'CPF_FINAL',
    'FONTE_TEL', 'TEL_FINAL',
    'DECISAO', 'OBSERVACAO'
]

# Cria apenas DECISAO e OBSERVACAO se não existirem
for col in ['DECISAO', 'OBSERVACAO']:
    if col not in df_revisao.columns:
        df_revisao[col] = ''

colunas_exportar = [c for c in colunas_organizadas if c in df_revisao.columns]
df_revisao[colunas_exportar].to_excel(
    f'{PASTA_OUTPUT}/PLANILHA_REVISAO_ENRIQUECIDA.xlsx', index=False)

# ==============================================
# ADICIONAR VALIDAÇÃO DE DADOS E FÓRMULAS DINÂMICAS
# ==============================================
print("\n🔒 Adicionando automações na planilha (Dropdowns e Fórmulas)...")

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = load_workbook(f'{PASTA_OUTPUT}/PLANILHA_REVISAO_ENRIQUECIDA.xlsx')
ws = wb.active

# 1. Cria as duas regras de validação (Dropdowns)
dv_decisao = DataValidation(type="list", formula1='"CONFIRMADO,REJEITADO"', allow_blank=True)
dv_fonte = DataValidation(type="list", formula1='"NENHUM,PESQUISA,IPTU,SAUDE"', allow_blank=True)

ws.add_data_validation(dv_decisao)
ws.add_data_validation(dv_fonte)

# 2. Mapeia e aplica os dropdowns
letras = {}
for col_idx, cell in enumerate(ws[1], 1):
    col_letra = get_column_letter(col_idx)
    nome_coluna = cell.value
    letras[nome_coluna] = col_letra
    
    # Aplica o dropdown de CONFIRMADO/REJEITADO na coluna DECISAO
    if nome_coluna == 'DECISAO':
        dv_decisao.add(f'{col_letra}2:{col_letra}{ws.max_row}')
        
    # Aplica o dropdown das fontes nas colunas FONTE_
    elif nome_coluna in ['FONTE_NOME', 'FONTE_CPF', 'FONTE_TEL']:
        dv_fonte.add(f'{col_letra}2:{col_letra}{ws.max_row}')

# 3. Injeta as fórmulas
for r in range(2, ws.max_row + 1):
    f_nome = f'=IF({letras["FONTE_NOME"]}{r}="PESQUISA", {letras["NOME_PESQUISA"]}{r}, IF({letras["FONTE_NOME"]}{r}="IPTU", {letras["NOME_IPTU"]}{r}, IF({letras["FONTE_NOME"]}{r}="SAUDE", {letras["NOME_SAUDE"]}{r}, "")))'
    ws[f'{letras["NOME_FINAL"]}{r}'] = f_nome

    f_cpf = f'=IF({letras["FONTE_CPF"]}{r}="PESQUISA", {letras["CPF_PESQUISA"]}{r}, IF({letras["FONTE_CPF"]}{r}="IPTU", {letras["CPF_IPTU"]}{r}, IF({letras["FONTE_CPF"]}{r}="SAUDE", {letras["CPF_SAUDE"]}{r}, "")))'
    ws[f'{letras["CPF_FINAL"]}{r}'] = f_cpf

    f_tel = f'=IF({letras["FONTE_TEL"]}{r}="PESQUISA", {letras["TEL_PESQUISA"]}{r}, IF({letras["FONTE_TEL"]}{r}="IPTU", {letras["TEL_IPTU"]}{r}, IF({letras["FONTE_TEL"]}{r}="SAUDE", {letras["TEL_SAUDE"]}{r}, "")))'
    ws[f'{letras["TEL_FINAL"]}{r}'] = f_tel

wb.save(f'{PASTA_OUTPUT}/PLANILHA_REVISAO_ENRIQUECIDA.xlsx')
print("   ✅ Fórmulas e Dropdowns inteligentes aplicados com sucesso!")
